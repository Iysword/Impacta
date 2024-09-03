import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import pandas as pd
import requests
import re
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import shutil
from datetime import datetime
from openpyxl.styles import PatternFill
import webbrowser
import tkinter.messagebox as messagebox
from tkinter import Text

def upload_file():
    file_path = filedialog.askopenfilename(filetypes=[("Planilhas", "*.xlsx;*.xls")])
    if file_path:
        df = pd.read_excel(file_path)
        if 'Master Waybill' and 'Controle CCT' in df.columns:
            data_master_waybill = df['Master Waybill'].tolist()
            data_waybill_number = df['Waybill Number'].tolist()
            data_controle_cct = df['Controle CCT'].tolist()
            data = list(zip(data_master_waybill, data_waybill_number, data_controle_cct))
            return data
    return [("A coluna 'Master Waybill' ou 'Controle CCT' não foram encontradas na planilha selecionada.", "", "")]

def update_treeview():
    tree.delete(*tree.get_children())
    for i, item in enumerate(text_to_display):
        tree.insert("", "end", values=(item[0], item[1], item[2]), tags=(i,))

def on_double_click(event):
    item = tree.selection()
    if item:
        tags = tree.item(item, "tags")
        if tags:
            tag = tags[0]
            current_tags = tree.item(item, "tags")
            if "selected" in current_tags:
                tree.item(item, tags=())
            else:
                tree.item(item, tags=("selected",))
                tree.tag_configure("selected", background="light blue")

    update_selected_items_counter()

def select_all_rows():
    checkbox_state = select_all_var.get()

    for i in tree.get_children():
        tree.item(i, tags=())  
    if checkbox_state:
        for i in tree.get_children():
            tree.item(i, tags=("selected",))
            tree.tag_configure("selected", background="light blue")

    update_selected_items_counter()

def select_same_master_waybill():
    item = tree.selection()
    if item:
        master_waybill = tree.item(item, "values")[0]
        for i in tree.get_children():
            if tree.item(i, "values")[0] == master_waybill:
                tree.item(i, tags=("selected",))
                tree.tag_configure("selected", background="light blue")

        update_selected_items_counter()

def on_right_click(event):
    item = tree.identify_row(event.y)
    tree.selection_set(item)  
    menu.post(event.x_root, event.y_root)

def update_selected_items_counter():
    selected_items = [item for item in tree.get_children() if "selected" in tree.item(item, "tags")]
    counter_label.config(text=f"Selecionados: {len(selected_items)}")

def update_widgets():
    global text_to_display
    text_to_display = upload_file()
    update_treeview()

def show_log_files():
    log_folder = "Log"

    # Certifique-se de que a pasta "Log" exista. Se não existir, crie-a.
    os.makedirs(log_folder, exist_ok=True)

    log_files = [f for f in os.listdir(log_folder) if f.endswith(".xlsx")]

    # Criar uma nova janela para exibir os arquivos de log
    log_window = tk.Toplevel(root)
    log_window.title("Arquivos de Log")

    # Criar uma treeview na nova janela com duas colunas
    log_tree = ttk.Treeview(log_window, columns=("Log", "Status"), show="headings")
    log_tree.heading("Log", text="Log")
    log_tree.heading("Status", text="Status")

    # Preencher a treeview com os logs e status
    for log_file in log_files:
        status = "Transmitido" if log_file.endswith("#.xlsx") else "Pendente de envio"
        log_tree.insert("", "end", values=(log_file, status))

    log_tree.pack(padx=10, pady=10)

    def open_selected_log():
        selected_log_file = log_tree.item(log_tree.selection(), "values")[0]
        if selected_log_file:
            selected_log_path = os.path.join(log_folder, selected_log_file)
            os.startfile(selected_log_path)

    def open_transmit_link():
        transmit_link = "https://plataforma.logcomex.io/signIn/"
        webbrowser.open(transmit_link)

    def delete_selected_log():
        selected_item = log_tree.selection()
        if selected_item:
            selected_log_file = log_tree.item(selected_item, "values")[0]
            if selected_log_file:
                selected_log_path = os.path.join(log_folder, selected_log_file)
                os.remove(selected_log_path)
                log_tree.delete(selected_item)

    # Adicionar botão para abrir o arquivo de log selecionado
    open_log_button = tk.Button(log_window, text="Abrir Log Selecionado", command=open_selected_log)
    open_log_button.pack(pady=10)

    # Adicionar menu de contexto para excluir o log selecionado
    context_menu = tk.Menu(log_window, tearoff=0)
    context_menu.add_command(label="Excluir Log Selecionado", command=delete_selected_log)

    transmit_button = tk.Button(log_window, text="Transmitir", command=open_transmit_link)
    transmit_button.pack(pady=10)

    def show_context_menu(event):
        context_menu.post(event.x_root, event.y_root)

    # Vincular o menu de contexto à treeview
    log_tree.bind("<Button-3>", show_context_menu)
def create_buttons():
    global upload_button, select_all_checkbox, select_all_var, counter_label, menu, store_button, selected_waybills

    upload_button = tk.Button(root, text="Upload de Planilha", command=update_widgets)
    upload_button.grid(row=0, column=0, pady=(10, 0), sticky="w", padx=(10, 0))

    select_all_var = tk.BooleanVar()
    select_all_checkbox = tk.Checkbutton(root, text="Selecionar Tudo", variable=select_all_var, command=select_all_rows)
    select_all_checkbox.grid(row=2, column=0, pady=(10, 0), sticky="w", padx=(10, 0))

    counter_label = tk.Label(root, text="Selecionados: 0")
    counter_label.grid(row=2, column=0, pady=(10, 0), sticky="n", padx=(0, 0))

    menu = tk.Menu(root, tearoff=0)
    menu.add_command(label="Selecionar Mesmo Master Waybill", command=select_same_master_waybill)

    store_button = tk.Button(root, text="Lançar", command=lançar)
    store_button.grid(row=0, column=0, pady=(10, 0), sticky="e", padx=(10, 0))

    selected_waybills = []
 
selected_waybills = []

def lançar():
    global selected_waybills
    selected_waybills = [tree.item(item, "values")[1] for item in tree.get_children() if "selected" in tree.item(item, "tags")]
    url = "https://37339.magayacloud.com/api/Invoke?Handler=CSSoapService"

    headers = {
        "Content-Type": "text/xml; charset=utf-8",
    }

    StartSession = """
    <s:Envelope xmlns:s="http://schemas.xmlsoap.org/soap/envelope/">
        <s:Body s:encodingStyle="http://schemas.xmlsoap.org/soap/encoding/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">
            <q1:StartSession xmlns:q1="urn:CSSoapService">
                <user xsi:type="xsd:string">api</user>
                <pass xsi:type="xsd:string">M@g@y@33166</pass>
            </q1:StartSession>
        </s:Body>
    </s:Envelope>
    """


    response_SS = requests.post(url, data=StartSession, headers=headers)
    if response_SS.status_code == 200:
  
        root = ET.fromstring(response_SS.text)
        key = root.find(".//access_key").text
        access_key = key
    else:    
        print("Falha na requisição. Código de status:", response_SS.status_code)  
    excel_file_path = "Planilha Padrão CCT Impo Aéreo.xlsx"
    book = openpyxl.load_workbook(excel_file_path)

 
    sheet = book['2. Cadastro (XFZB + XFHL)']

    row_index = 5
    for house in selected_waybills:  
        GetTransRange = f"""<s:Envelope xmlns:s="http://schemas.xmlsoap.org/soap/envelope/">
                <s:Body s:encodingStyle="http://schemas.xmlsoap.org/soap/encoding/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="https://schema.magaya.net/Core/V1/Shipment.xsd">
                    <GetTransaction xmlns="urn:CSSoapService">
                        <access_key>{access_key}</access_key>
                        <type>SH</type>
                        <flags>1</flags>
                        <number>{house}</number>                   
                    </GetTransaction>
                </s:Body>
            </s:Envelope>"""
    
        response_GT = requests.post(url, data=GetTransRange, headers=headers)
        if response_GT.status_code != 200:
            print("Falha:", response_GT.status_code)
      

        GT_Data = response_GT.text
    
        root_GT = ET.fromstring(GT_Data)
     
        trans_xml = root_GT.find('.//trans_xml').text
    
        if trans_xml is not None:
            internal_root_wr = ET.fromstring(trans_xml)

        else:
            print("Erro: XML vazio ou não encontrado na resposta da solicitação HTTP")
        origin_port_element = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}OriginPort')
        origin = origin_port_element.get('Code')
        DestinationPort = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}DestinationPort')
        DestinationAgent = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}IssuedBy/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}Country').text
        Destin = DestinationPort.get('Code')
        shipment_number = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Number')
        if shipment_number is not None:
            shipment_number = shipment_number.text.replace('-', '')
        shipment_pieces = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}TotalPieces')
        if shipment_pieces is not None:
            shipment_pieces = shipment_pieces.text
        shipper_name = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}ShipperName')
        if shipper_name is not None:
            shipper_name = shipper_name.text
        consignee_name = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}ConsigneeName')
        if consignee_name is not None:
            consignee_name = consignee_name.text
        print(house)
        if DestinationAgent == 'United States' :
            localdeemissão = 'Doral'
        else:
            localdeemissão = 'Paris'
        carrier_name = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}CarrierName')
        ttalweigtht = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}TotalWeight')
        if DestinationAgent == 'United States':
            expedidor = 'JOSÉ FERREIRA DE FREITAS'
        else:
            expedidor = 'DIOGO GUERRA'    
            
        if ttalweigtht is not None:
            ttalweigtht = float(ttalweigtht.text) if ttalweigtht.text else 0.00 #análise
            ttalweigtht = "{:.2f}".format(ttalweigtht)
        Cfields_elements = internal_root_wr.findall(".//{http://www.magaya.com/XMLSchema/V1}CustomFields/{http://www.magaya.com/XMLSchema/V1}CustomField")
        for elem in Cfields_elements:
            if elem.find(".//{http://www.magaya.com/XMLSchema/V1}CustomFieldDefinition/{http://www.magaya.com/XMLSchema/V1}InternalName").text == 'dta_rm_carga':
                Cfields = elem.findall(".//{http://www.magaya.com/XMLSchema/V1}Value")
                Cfields = Cfields[1].text
                break
        else:
            Cfields = None
        currency = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Charges/{http://www.magaya.com/XMLSchema/V1}Charge/{http://www.magaya.com/XMLSchema/V1}Currency')
        charges = internal_root_wr.findall('.//{http://www.magaya.com/XMLSchema/V1}Charges/{http://www.magaya.com/XMLSchema/V1}Charge')
        sum = 0
        totalcollect = 0
        totalprepaid = 0
        for charge in charges:
            amount = charge.find('.//{http://www.magaya.com/XMLSchema/V1}AmountInCurrency').text
        am = float(amount)
        sum = sum + am 
        soma = 0.0
        soma1 = 0.0
        for charge in charges:
            amount_element = charge.find('.//{http://www.magaya.com/XMLSchema/V1}AmountInCurrency')
            description = charge.find('.//{http://www.magaya.com/XMLSchema/V1}ChargeDefinition/{http://www.magaya.com/XMLSchema/V1}Description').text
            if amount_element is not None:
                amount = amount_element.text
                if 'FREIGHT' in description:
                    pesotot = float(amount)
                    pesotot = round(pesotot, 2)
                    soma1 += pesotot
                try:
                    am1 = float(amount)
                    am1 = round(am1, 2)
                    soma += am1
                except ValueError:
                    
                    pass
            else:
                
                pass    
        cget = currency.get('Code')
        moedaprepaid = cget
        moedacollect = cget
        PesoCobrança = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}ChargeableWeight')
        if PesoCobrança is not None:
            PesoCobrança = float(PesoCobrança.text) if PesoCobrança.text else 0.00
        PesoCubado = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}VolumeWeight')
        if PesoCubado is not None:
            PesoCubado = float(PesoCubado.text) if PesoCubado.text else 0.00
            PesoCubado = "{:.2f}".format(PesoCubado)
        Description = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}DescriptionOfGoods')
        CNPJAgente = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}DestinationAgent/{http://www.magaya.com/XMLSchema/V1}ExporterID')
        if CNPJAgente is not None:
            CNPJAgente = CNPJAgente.text
            CNPJAgente = re.sub(r'[^0-9]', '', CNPJAgente)
        else:
            CNPJAgente = 'vazio'
        cneecity = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}City')
        if cneecity is not None:
            cneecity = cneecity.text
        else:
            cneecity ='vazio'
        cneerua = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}Street')
        if cneerua is not None:
            cneerua = cneerua.text
        else:
            cneerua = 'vazio'
        cneecountry = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}Country')
        if cneecountry is not None:
            cnecontrycode = cneecountry.get('Code')
        else:
            cnecontrycode = 'vazio'
        Shipperrua = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Shipper/{http://www.magaya.com/XMLSchema/V1}Address//{http://www.magaya.com/XMLSchema/V1}Street')
        if Shipperrua is not None:
            Shipperrua = Shipperrua.text
        else:
            Shipperrua = 'vazio'
        Shippercity = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Shipper/{http://www.magaya.com/XMLSchema/V1}Address//{http://www.magaya.com/XMLSchema/V1}City')
        if Shippercity is not None:
            Shippercity = Shippercity.text
        else:
            Shippercity = 'vazio'
        shippercountry = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Shipper/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}Country')
        cneeZipcode = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}ZipCode')
        if cneeZipcode is not None:
            cneeZipcode = cneeZipcode.text
            cneeZipcode = re.sub(r'[^0-9]', '', cneeZipcode)
        else:
            cneeZipcode = 'vazio'
        Shipperzipcode = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Shipper/{http://www.magaya.com/XMLSchema/V1}Address/{http://www.magaya.com/XMLSchema/V1}ZipCode')
        if Shipperzipcode  is not None:
            Shipperzipcode = Shipperzipcode.text
            Shipperzipcode = re.sub(r'[^0-9]', '', Shipperzipcode)
        else:
            Shipperzipcode = 'vazio'
        if shippercountry is not None:    
            shippercountrycode = shippercountry.get('Code')
            if shippercountrycode is None:
                shippercountrycode = 'vazio'
        Master = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}MasterWayBillNumber').text
        pp = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}IsPrepaid')
        if pp is not None:
            pp = pp.text
            if pp == 'true':
                pp = 'PREPAID'
                totalprepaid = soma
            else:
                pp = 'COLLECT'
                totalcollect = soma
        airports_data = {
        'GRU': {'code': '0817600', 'number': '8911101', 'value': '500'},
        'VCP': {'code': '0817700', 'number': '8921101', 'value': '011'},
        'CWB': {'code': '0915200', 'number': '9991101', 'value': '015'},
        'GIG': {'code': '0717700', 'number': '7911101', 'value': '010'},
        'FOR': {'code': '0317700', 'number': '3921101', 'value': '001'} }
        cneetype = 'CNPJ'
        cneeCNPJ = internal_root_wr.find('.//{http://www.magaya.com/XMLSchema/V1}Consignee/{http://www.magaya.com/XMLSchema/V1}ExporterID')
        
        if cneeCNPJ is not None:
            cneeCNPJ = cneeCNPJ.text
            cneeCNPJ = re.sub(r'[^0-9]', '', cneeCNPJ)
        
        if Description is not None:
            DESC = Description.text
            if 'TREATED AND CERTIFIED' in DESC:
                madeira = 'SIM'
            else:
                madeira = 'NÃO' 
            if DESC is not None:
                DESC = DESC.splitlines()
                DESC1 = DESC[0]
                DESC2 = DESC[1]
                DESC = f'{DESC1}\n{DESC2}'
                
        else:
            DESC = 'vazio'
        if Cfields == 'Yes':
            code = internal_root_wr.findall(".//{http://www.magaya.com/XMLSchema/V1}CustomFields/{http://www.magaya.com/XMLSchema/V1}CustomField")
            for elemcode in code:
                if elemcode.find(".//{http://www.magaya.com/XMLSchema/V1}CustomFieldDefinition/{http://www.magaya.com/XMLSchema/V1}InternalName").text == 'recinto_aduaneiro':
                    Aduacode = elemcode.find(".//{http://www.magaya.com/XMLSchema/V1}Value").text
                    Aduacode = Aduacode[:7]
        else:            
            if Destin in airports_data:
                Aduacode = airports_data[Destin]['code']
            else:
                Aduacode = None
        Manuseio = []
        Manu = internal_root_wr.findall(".//{http://www.magaya.com/XMLSchema/V1}CustomFields/{http://www.magaya.com/XMLSchema/V1}CustomField")
        for manucode in Manu:
            if manucode.find(".//{http://www.magaya.com/XMLSchema/V1}CustomFieldDefinition/{http://www.magaya.com/XMLSchema/V1}Category") is not None:
                if manucode.find(".//{http://www.magaya.com/XMLSchema/V1}CustomFieldDefinition/{http://www.magaya.com/XMLSchema/V1}Category").text ==  'CCT - Manuseio':
                    carga = manucode.find(".//{http://www.magaya.com/XMLSchema/V1}Value").text
                    nomemanu = manucode.find(".//{http://www.magaya.com/XMLSchema/V1}CustomFieldDefinition/{http://www.magaya.com/XMLSchema/V1}DisplayName").text
                    if carga == 'true':
                        Manuseio.append(nomemanu)
        manuseiodecarga = ','.join(Manuseio)
                
        creationdate = internal_root_wr.find(".//{http://www.magaya.com/XMLSchema/V1}CreatedOn").text        
        GetTransMaster = f"""<s:Envelope xmlns:s="http://schemas.xmlsoap.org/soap/envelope/">
                <s:Body s:encodingStyle="http://schemas.xmlsoap.org/soap/encoding/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="https://schema.magaya.net/Core/V1/Shipment.xsd">
                    <GetTransaction xmlns="urn:CSSoapService">
                        <access_key>{access_key}</access_key>
                        <type>SH</type>
                        <flags>1</flags>
                        <number>{Master}</number>                   
                    </GetTransaction>
                </s:Body>
            </s:Envelope>"""

        response_Master = requests.post(url, data=GetTransMaster, headers=headers)
        if response_Master.status_code != 200:
            print("Falha:", response_Master.status_code)

        Master_Data = response_Master.text
        root_Master = ET.fromstring(Master_Data)
        trans_xmlm = root_Master.find('.//trans_xml').text
        internal_root_wrm = ET.fromstring(trans_xmlm)

        if totalcollect != 0.00:
            otherchargescollect = totalcollect - soma1
        else:
            otherchargescollect = 0.00
        if totalprepaid != 0.00:
            otherchargespp = totalprepaid - soma1
        else:
            otherchargespp = 0.00
        if expedidor == 'JOSÉ FERREIRA DE FREITAS':
            embarcador = 'DUX FORWARDING'
            embarcadorcountry = 'US'
            embarcadoradreess = '10206 NW 19TH ST STE 100'
            embarcadorcity = 'DORAL'
        else: 
            embarcador = 'DUX LOGISTIQUE'
            embarcadorcountry = 'FR'
            embarcadoradreess =  '42 Rue de Maubeuge, Charles de Gaulle'
            embarcadorcity = 'CHARLES DE GAULLE'
            

        MasterPieces = internal_root_wrm.find('.//{http://www.magaya.com/XMLSchema/V1}TotalPieces').text
        MasterWeight = internal_root_wrm.find('.//{http://www.magaya.com/XMLSchema/V1}TotalWeight').text
        MasterWeight = float(MasterWeight)
        creationdate = datetime.fromisoformat(creationdate)
        creationdate = creationdate.strftime("%Y%m%d")
        MasterWeight = "{:.2f}".format(MasterWeight)
        soma1 = "{:.2f}".format(float(soma1))
        soma = "{:.2f}".format(float(soma))
        PesoCobrança = "{:.2f}".format(float(PesoCobrança))
        totalcollect = "{:.2f}".format(float(totalcollect))
        totalprepaid = "{:.2f}".format(float(totalprepaid))
        otherchargescollect = "{:.2f}".format(float(otherchargescollect))
        otherchargespp = "{:.2f}".format(float(otherchargespp))
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        lista = ['C', shipment_number, origin, Destin, shipment_pieces, ttalweigtht, pp, cget, soma1, PesoCubado, PesoCobrança, 'KGM', DESC, moedaprepaid, totalprepaid, moedacollect, totalcollect, CNPJAgente, consignee_name, cneerua, cneecity, cnecontrycode, cneeZipcode, shipper_name, Shipperrua, Shippercity, shippercountrycode, Shipperzipcode, Aduacode, Master, MasterWeight, MasterPieces, origin, Destin, None, cneetype, cneeCNPJ, None, expedidor ,None , None, None, madeira, None, manuseiodecarga, None, creationdate, localdeemissão, 'MTQ', otherchargespp, otherchargescollect, 0.00, 0.00,None , DESC, embarcador, embarcadoradreess, embarcadorcity,  embarcadorcountry,None,'DUX AGENCIAMENTO DE CARGAS LTDA']
        for col_index, value in enumerate(lista, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if value == 'vazio':
                cell.fill = red_fill
        row_index += 1


    book.save(excel_file_path)
    log_folder = "Log"
    
    # Obtém a data e hora atual
    current_datetime = datetime.now().strftime("%d_%m_%H_%M")
    
    # Formata o nome do arquivo de log
    log_file_name = f"Shipment list-{current_datetime}.xlsx"
    log_excel_file_path = os.path.join(log_folder, log_file_name)

    try:
        # Certifique-se de que a pasta "Log" exista. Se não existir, crie-a.
        os.makedirs(log_folder, exist_ok=True)
        
        # Crie uma cópia da planilha na pasta "Log"
        shutil.copy(excel_file_path, log_excel_file_path)

        print(f"Cópia da planilha criada com sucesso na pasta 'Log' como {log_file_name}")

        # Abrir a planilha original para edição
        book = openpyxl.load_workbook(excel_file_path)
        sheet = book.active

        # Deletar as informações da linha 5 pra baixo
        sheet.delete_rows(5, sheet.max_row)

        # Salvar as mudanças na planilha original
        book.save(excel_file_path)

    except Exception as e:
        print(f"Erro ao criar cópia da planilha na pasta 'Log': {e}")
        book.save(excel_file_path)

    # Adicionando código para abrir a planilha na tela
    os.startfile(log_excel_file_path)
   
    

    
root = tk.Tk()
root.title("Upload de Planilha")
root.geometry("650x380")  

frame = ttk.Frame(root, padding=(10, 0, 0, 0))  
frame.grid(row=1, column=0, pady=(20, 0), sticky="nsew")
tree = ttk.Treeview(frame, columns=("Master Waybill", "Waybill Number", "Controle CCT"), show="headings", selectmode="extended")
tree.heading("Master Waybill", text="Master Waybill")
tree.heading("Waybill Number", text="Waybill Number")
tree.heading("Controle CCT", text="Status")


style = ttk.Style()
style.configure("Treeview.Heading", anchor="center")
style.configure("Treeview", rowheight=25)
style.configure("Treeview.Treeitem", padding=(0, 0, 0, 0)) 


tree.column("#0", stretch=tk.NO, width=1)
tree.column("Master Waybill", anchor="center", width=200)  
tree.column("Waybill Number", anchor="center", width=200)  
tree.column("Controle CCT", anchor="center", width=200)   

tree.grid(row=0, column=0, pady=(0, 0), sticky="nsew")

scrollbar = ttk.Scrollbar(frame, command=tree.yview)
scrollbar.grid(row=0, column=1, pady=(0, 0), sticky='ns')
tree.configure(yscrollcommand=scrollbar.set)

tree.bind("<Double-1>", on_double_click)

tree.bind("<Button-3>", on_right_click)

create_buttons()
log_button = tk.Button(root, text="Log", command=show_log_files)
log_button.grid(row=0, column=0, pady=(10, 0), sticky="s", padx=(10, 5))
frame.grid_rowconfigure(0, weight=1)
frame.grid_columnconfigure(0, weight=1)

root.mainloop()
